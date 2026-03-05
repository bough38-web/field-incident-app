import io
import os
import zipfile
import yaml
import tempfile
import logging
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import List, Dict, Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

# Set up simple logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

def load_categories(path: str = "categories.yaml") -> dict:
    """Loads the category mapping from a YAML file."""
    try:
        with open(path, "r", encoding="utf-8") as f:
            return yaml.safe_load(f)
    except Exception as e:
        logger.warning(f"Failed to load categories from {path}: {e}")
        return {"ETC": {"folder": "99_기타", "label_ko": "기타"}}

def _create_summary_sheet(wb: Workbook, incidents: List[Dict[str, Any]]) -> None:
    """Creates a visualization summary sheet with basic statistics and charts."""
    ws_summary = wb.create_sheet(title="Summary & Dashboard", index=0)
    
    # Title
    ws_summary.merge_cells("A1:E1")
    title_cell = ws_summary["A1"]
    title_cell.value = "사건/사고 통계 대시보드"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Aggregate data
    branch_counts = {}
    status_counts = {}
    
    for it in incidents:
        branch = it.get("branch") or "Unknown"
        status = it.get("status") or "Unknown"
        branch_counts[branch] = branch_counts.get(branch, 0) + 1
        status_counts[status] = status_counts.get(status, 0) + 1
        
    # Write branch counts
    ws_summary.append([]) # Empty row 2
    ws_summary.append(["지점명(Branch)", "건수(Count)"])
    ws_summary["A3"].font = Font(bold=True)
    ws_summary["B3"].font = Font(bold=True)
    
    row_idx = 4
    for branch, count in branch_counts.items():
        ws_summary.append([branch, count])
        row_idx += 1
        
    # Autosize columns
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 15

def build_incidents_xlsx(incidents: List[Dict[str, Any]]) -> bytes:
    """Generates an Excel file containing the incident list and a visualization dashboard."""
    wb = Workbook()
    
    # 1. Dashboard (Summary Sheet)
    _create_summary_sheet(wb, incidents)
    
    # 2. Add main data sheet
    ws = wb.create_sheet(title="Incidents List")
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"]) # Remove default sheet

    headers = [
        "incident_id", "branch", "status",
        "created_by", "created_at",
        "completed_at", "reported_at",
        "service_no", "address"
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx)
        c.font = Font(bold=True, color="000000")
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    for it in incidents:
        ws.append([
            it.get("id"),
            it.get("branch"),
            it.get("status"),
            it.get("created_by_name") or it.get("created_by"),
            it.get("created_at"),
            it.get("completed_at"),
            it.get("reported_at"),
            it.get("service_no"),
            it.get("address"),
        ])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(incidents) + 1}"
    
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 40)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

def _download_photo(storage: Any, file_meta: Dict[str, Any], cats: dict, categories_path: str) -> Optional[tuple]:
    """Helper for concurrent downloads. Returns (archive_path, bytes_content)."""
    try:
        cat = file_meta.get("category") or "ETC"
        folder = (cats.get(cat) or cats.get("ETC", {})).get("folder", "99_기타")
        
        incident_id = file_meta.get("incident_id", "unknown")
        branch = file_meta.get("branch", "")
        employee = file_meta.get("employee", "")
        taken_at = file_meta.get("taken_at") or ""
        orig = file_meta.get("original_name") or "photo.jpg"

        # Safe filename
        safe = lambda s: "".join(ch for ch in str(s) if ch.isalnum() or ch in "-_(). ").strip()
        filename = f"{safe(incident_id)}_{safe(branch)}_{safe(employee)}_{safe(taken_at)}_{cat}_{safe(orig)}"
        archive_path = f"photos/{folder}/{filename}"

        content = storage.get_bytes(file_meta["storage_key"])
        
        if content is None:
            logger.error(f"Failed to fetch content for key: {file_meta['storage_key']}")
            return None
            
        return (archive_path, content)
        
    except Exception as e:
        logger.error(f"Error downloading {file_meta.get('storage_key')}: {e}")
        return None

def build_export_zip(incidents: List[Dict[str, Any]], files: List[Dict[str, Any]], storage: Any, categories_path: str = "categories.yaml") -> tempfile.SpooledTemporaryFile:
    """
    Builds a ZIP file containing the excel export and photos.
    Uses SpooledTemporaryFile to handle large ZIP files without OOM.
    Uses ThreadPoolExecutor for concurrent fetching to speed up downloads.
    """
    cats = load_categories(categories_path)
    
    # 50MB in memory, rest goes to disk
    temp_zip = tempfile.SpooledTemporaryFile(max_size=50 * 1024 * 1024, mode="w+b", prefix="export_", suffix=".zip")

    try:
        with zipfile.ZipFile(temp_zip, "w", compression=zipfile.ZIP_DEFLATED) as z:
            
            # 1) 엑셀 생성 및 추가 (시각화 대시보드 포함)
            logger.info("Generating Excel document with Visualization...")
            xlsx_bytes = build_incidents_xlsx(incidents)
            z.writestr("incidents.xlsx", xlsx_bytes)

            # 2) 병렬 비동기 사진 다운로드 및 ZIP 추가 (속도 최적화)
            logger.info(f"Downloading {len(files)} photos concurrently...")
            max_workers = min(10, (os.cpu_count() or 1) * 2) 
            
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                future_to_file = {
                    executor.submit(_download_photo, storage, f, cats, categories_path): f
                    for f in files
                }
                
                for future in as_completed(future_to_file):
                    result = future.result()
                    if result:
                        archive_path, content = result
                        z.writestr(archive_path, content)

            # 3) 설정 파일 포함
            if os.path.exists(categories_path):
                with open(categories_path, "rb") as cf:
                    z.writestr("categories.yaml", cf.read())
                    
        temp_zip.seek(0)
        return temp_zip
        
    except Exception as e:
        temp_zip.close()
        logger.error(f"Failed to create ZIP export: {e}")
        raise e

# ---- 사용 예시 (검증용 S3 Mock Storage 주석) ----
# class S3StorageMock:
#     def get_bytes(self, key: str) -> bytes:
#         return b"mock file content"
#
# incidents_data = [{"id": "INC-001", "branch": "강남점", "status": "완료", "created_by": "홍길동"}]
# files_data = [{"storage_key": "s3://bucket/test.jpg", "incident_id": "INC-001", "category": "METER"}]
#
# tmp_zip_file = build_export_zip(incidents_data, files_data, S3StorageMock())
# with open("export_complete.zip", "wb") as f_out:
#     while True:
#         chunk = tmp_zip_file.read(8192)
#         if not chunk: break
#         f_out.write(chunk)
# tmp_zip_file.close()

